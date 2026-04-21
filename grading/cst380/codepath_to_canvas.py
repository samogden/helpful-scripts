#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import inspect
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from lms_interface.canvas_interface import CanvasInterface
import yaml
from thefuzz import fuzz
from thefuzz import process


POINTS_POSSIBLE_LABEL = "    Points Possible"
LOS_ANGELES = ZoneInfo("America/Los_Angeles")


@dataclass(frozen=True)
class ScoreConfig:
  base_points: float
  stretch_points: float
  ignore_points: float
  stretch_weight: float
  raw_output_points: float
  canvas_value: float
  canvas_scale: float
  effective_base_points: float
  effective_stretch_points: float
  effective_total_points: float


@dataclass(frozen=True)
class MatchSuggestion:
  canvas_name: str
  score: int


@dataclass(frozen=True)
class ScoreBreakdown:
  raw_score: float
  adjusted_raw_score: float
  base_earned: float
  stretch_earned: float
  weighted_total: float
  canvas_score: float


@dataclass(frozen=True)
class SubmissionDecision:
  action: str
  due_at: datetime | None
  submitted_at: datetime | None = None
  seconds_late: int | None = None


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
  parser = argparse.ArgumentParser(
    description="Push CodePath grades to Canvas in batch using assignments.yaml."
  )
  parser.add_argument("--in", dest="codepath_csv", help=argparse.SUPPRESS)
  parser.add_argument(
    "--out",
    help=argparse.SUPPRESS,
  )
  parser.add_argument(
    "--assignments",
    help="YAML file with per-assignment settings for batch processing.",
  )
  parser.add_argument(
    "--data-dir",
    help="Directory containing codepath-<assignment>.csv and canvas-<assignment>.csv files.",
  )
  parser.add_argument(
    "--xls",
    help="Gradebook workbook to use as the source of assignment rows, with sheets like 'ASN - 1'.",
  )
  parser.add_argument(
    "--only-assignment",
    action="append",
    help="Limit batch processing to one or more assignment keys from --assignments.",
  )
  parser.add_argument("--canvas", help=argparse.SUPPRESS)
  parser.add_argument("--assignment-column", help=argparse.SUPPRESS)
  parser.add_argument(
    "--name-map",
    default="name_map.yaml",
    help="YAML file storing Canvas canonical names with CodePath aliases, like peer-eval's names.yaml. Defaults to ./name_map.yaml",
  )
  parser.add_argument(
    "--write-suggestions",
    help=argparse.SUPPRESS,
  )
  parser.add_argument(
    "--prompt-for-matches",
    action="store_true",
    help=argparse.SUPPRESS,
  )
  parser.add_argument("--base-points", type=float, help=argparse.SUPPRESS)
  parser.add_argument("--stretch-points", type=float, help=argparse.SUPPRESS)
  parser.add_argument(
    "--ignore-points",
    default=0.0,
    type=float,
    help=argparse.SUPPRESS,
  )
  parser.add_argument(
    "--stretch-weight",
    type=float,
    default=0.5,
    help=argparse.SUPPRESS,
  )
  parser.add_argument(
    "--canvas-value",
    type=float,
    default=None,
    help=argparse.SUPPRESS,
  )
  parser.add_argument(
    "--leave-not-graded-blank",
    action="store_true",
    help=argparse.SUPPRESS,
  )
  parser.add_argument(
    "--missing-as-zero",
    action="store_true",
    help=argparse.SUPPRESS,
  )
  parser.add_argument(
    "--auto-match-threshold",
    default=96,
    type=int,
    help=argparse.SUPPRESS,
  )
  parser.add_argument(
    "--auto-match-gap",
    default=4,
    type=int,
    help=argparse.SUPPRESS,
  )
  parser.add_argument(
    "--suggestion-count",
    default=5,
    type=int,
    help=argparse.SUPPRESS,
  )
  parser.add_argument("--verbose", action="store_true", help=argparse.SUPPRESS)
  parser.add_argument(
    "--prod",
    action="store_true",
    help="Push grades to Canvas prod. Default behavior uses Canvas dev.",
  )
  parser.add_argument(
    "--strict-deadlines",
    action="store_true",
    help="Use CodePath Updated instead of Submitted for deadline comparison when both exist.",
  )
  args = parser.parse_args(argv)

  if args.assignments and sys.stdin.isatty():
    args.prompt_for_matches = True

  if args.stretch_weight < 0:
    parser.error("--stretch-weight must be non-negative.")

  single_mode = bool(args.codepath_csv or args.canvas)
  batch_mode = bool(args.assignments or args.data_dir or args.xls)
  if single_mode and batch_mode:
    parser.error("Use either single-file mode (--in/--canvas) or batch mode (--assignments with --data-dir or --xls).")
  if not single_mode and not batch_mode:
    parser.error("Provide either --in/--canvas or --assignments with --data-dir or --xls.")

  if single_mode:
    if not args.codepath_csv or not args.canvas:
      parser.error("Single-file mode requires both --in and --canvas.")
    if args.base_points is None or args.stretch_points is None:
      parser.error("Single-file mode requires --base-points and --stretch-points.")
    if args.base_points < 0 or args.stretch_points < 0 or args.ignore_points < 0:
      parser.error("Point values must be non-negative.")
  else:
    if not args.assignments or (not args.data_dir and not args.xls):
      parser.error("Batch mode requires --assignments and either --data-dir or --xls.")
    if args.out:
      parser.error("--out is only supported in single-file mode.")

  return args


def parse_canvas_points_possible(
  canvas_rows: list[dict[str, str]],
  assignment_column: str,
) -> float | None:
  for row in canvas_rows:
    if row.get("Student", "") != POINTS_POSSIBLE_LABEL:
      continue
    raw_value = row.get(assignment_column, "").strip()
    if not raw_value:
      return None
    return float(raw_value)
  return None


def build_score_config(
  args: argparse.Namespace,
  canvas_rows: list[dict[str, str]],
  assignment_column: str,
) -> ScoreConfig:
  effective_total = max(args.base_points + args.stretch_points - args.ignore_points, 0.0)
  effective_base = min(args.base_points, effective_total)
  effective_stretch = max(effective_total - effective_base, 0.0)
  stretch_weight = args.stretch_weight
  raw_output_points = effective_base + effective_stretch * stretch_weight
  detected_canvas_value = parse_canvas_points_possible(canvas_rows, assignment_column)
  canvas_value = args.canvas_value if args.canvas_value is not None else detected_canvas_value
  if canvas_value is None:
    canvas_value = 100.0
  if canvas_value < 0:
    raise ValueError("--canvas-value must be non-negative.")

  canvas_scale = 0.0
  if raw_output_points > 0:
    canvas_scale = canvas_value / raw_output_points

  return ScoreConfig(
    base_points=args.base_points,
    stretch_points=args.stretch_points,
    ignore_points=args.ignore_points,
    stretch_weight=stretch_weight,
    raw_output_points=raw_output_points,
    canvas_value=canvas_value,
    canvas_scale=canvas_scale,
    effective_base_points=effective_base,
    effective_stretch_points=effective_stretch,
    effective_total_points=effective_total,
  )


def normalize_name(name: str) -> str:
  normalized = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
  return " ".join("".join(ch.lower() if ch.isalnum() else " " for ch in normalized).split())


def token_key(name: str) -> tuple[str, ...]:
  return tuple(sorted(normalize_name(name).split()))


def token_key_without_initials(name: str) -> tuple[str, ...]:
  return tuple(sorted(token for token in normalize_name(name).split() if len(token) > 1))


def read_codepath_rows(path: Path) -> list[dict[str, str]]:
  with path.open(newline="", encoding="utf-8-sig") as handle:
    reader = csv.DictReader(handle)
    return list(reader)


def normalize_sheet_text(value) -> str:
  if value is None:
    return ""
  if isinstance(value, float) and value.is_integer():
    return str(int(value))
  return str(value).strip()


def find_gradebook_workbook(data_dir: Path) -> Path | None:
  candidate_dirs: list[Path] = [Path.cwd()]
  if data_dir.is_dir() and data_dir != Path.cwd():
    candidate_dirs.append(data_dir)
  if data_dir.is_file() and data_dir.suffix.lower() == ".xlsx" and not data_dir.name.startswith("~$"):
    return data_dir

  candidates: list[Path] = []
  for directory in candidate_dirs:
    candidates.extend(
      sorted(
        path
        for path in directory.glob("*.xlsx")
        if not path.name.startswith("~$")
      )
    )

  if not candidates:
    return None

  preferred = [path for path in candidates if "gradebook" in path.name.lower()]
  if len(preferred) == 1:
    return preferred[0]
  if len(candidates) == 1:
    return candidates[0]
  raise ValueError(
    "Found multiple XLSX files; keep a single gradebook workbook in the working directory or pass the workbook path as --data-dir."
  )


def load_gradebook_assignment_rows(
  workbook_path: Path,
  assignment_name: str,
) -> tuple[list[dict[str, str]], str | None]:
  try:
    import openpyxl
  except ModuleNotFoundError as exc:
    raise ModuleNotFoundError(
      "openpyxl is required to read the gradebook workbook. Install requirements.txt first."
    ) from exc

  workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
  try:
    if assignment_name not in workbook.sheetnames:
      return [], f"Gradebook sheet {assignment_name!r} was not found in {workbook_path.name}."

    worksheet = workbook[assignment_name]
    row_iter = worksheet.iter_rows(values_only=True)
    header_found = False
    header_indexes: dict[str, int] = {}
    rows: list[dict[str, str]] = []
    seen_data = False
    empty_member_id_run = 0

    def get_value(values_row, header: str) -> str:
      index = header_indexes.get(header)
      if index is None or not values_row or index >= len(values_row):
        return ""
      return normalize_sheet_text(values_row[index])

    for row_index, values in enumerate(row_iter, start=1):
      if not header_found:
        if row_index > 20:
          break
        normalized_headers = {
          normalize_sheet_text(cell).strip().lower(): index
          for index, cell in enumerate(values or ())
          if normalize_sheet_text(cell).strip()
        }
        required = {"member id", "status", "full name", "feature score"}
        if required.issubset(normalized_headers):
          header_found = True
          header_indexes = dict(normalized_headers)
        continue

      member_id_index = header_indexes.get("member id", 0)
      member_id_value = values[member_id_index] if values and member_id_index < len(values) else None
      member_id = normalize_sheet_text(member_id_value)
      if not member_id:
        if seen_data:
          empty_member_id_run += 1
          if empty_member_id_run >= 25:
            break
        continue

      seen_data = True
      empty_member_id_run = 0
      row = {
        "Member ID": member_id,
        "Status": get_value(values, "status"),
        "Full Name": get_value(values, "full name"),
        "Feature Score": get_value(values, "feature score"),
        "Submitted": get_value(values, "submitted"),
        "Updated": get_value(values, "updated"),
      }
      if not row["Full Name"]:
        continue
      rows.append(row)

    if not header_found:
      raise ValueError(
        f"Could not find the gradebook header row in sheet {assignment_name!r} of {workbook_path.name}."
      )

    if not rows:
      return [], f"Gradebook sheet {assignment_name!r} contains no student rows yet."
    return rows, None
  finally:
    workbook.close()


def resolve_batch_assignment_input(
  *,
  assignment_name: str,
  data_dir: Path,
  gradebook_path: Path | None,
  prefer_gradebook: bool = False,
) -> tuple[Path | None, list[dict[str, str]] | None, str | None, str | None]:
  if gradebook_path is not None:
    if prefer_gradebook:
      gradebook_rows, skip_message = load_gradebook_assignment_rows(gradebook_path, assignment_name)
      return None, gradebook_rows, skip_message, None
  codepath_path = data_dir / f"codepath-{assignment_name}.csv"
  if codepath_path.exists():
    return codepath_path, None, None, None

  if gradebook_path is not None:
    gradebook_rows, skip_message = load_gradebook_assignment_rows(gradebook_path, assignment_name)
    return None, gradebook_rows, skip_message, None

  return None, None, None, f"Missing CodePath CSV for {assignment_name}: expected {codepath_path.name}"


def read_canvas_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
  with path.open(newline="", encoding="utf-8-sig") as handle:
    reader = csv.DictReader(handle)
    rows = list(reader)
    return reader.fieldnames or [], rows


def load_name_map(path: Path | None) -> dict[str, str]:
  if path is None or not path.exists():
    return {}

  loaded = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
  if isinstance(loaded, dict) and "CONFIRMED" in loaded:
    loaded = loaded["CONFIRMED"] or {}
  if isinstance(loaded, dict) and "matches" in loaded:
    loaded = loaded["matches"] or {}

  if not isinstance(loaded, dict):
    raise ValueError(f"Name map at {path} must be a mapping or contain a top-level 'matches' key.")

  if all(isinstance(value, list) for value in loaded.values()):
    mapping: dict[str, str] = {}
    for canonical_name, aliases in loaded.items():
      if canonical_name == "UNMATCHED":
        continue
      for alias in aliases:
        mapping[str(alias)] = str(canonical_name)
    return mapping

  return {str(source): str(target) for source, target in loaded.items() if target}


def group_aliases_by_canonical(mapping: dict[str, str]) -> dict[str, list[str]]:
  grouped: dict[str, list[str]] = {}
  for source_name, canonical_name in sorted(mapping.items(), key=lambda item: (item[1], item[0])):
    grouped.setdefault(canonical_name, []).append(source_name)
  return grouped


def save_name_map(
  path: Path,
  confirmed_mapping: dict[str, str],
  suggested_mapping: dict[str, str] | None = None,
  unmatched: list[str] | None = None,
) -> None:
  payload: dict[str, object] = {}
  confirmed_grouped = group_aliases_by_canonical(confirmed_mapping)
  if confirmed_grouped:
    payload["CONFIRMED"] = confirmed_grouped
  if suggested_mapping:
    payload["SUGGESTED"] = group_aliases_by_canonical(suggested_mapping)
  if unmatched:
    payload["UNMATCHED"] = sorted(unmatched)

  with path.open("w", encoding="utf-8") as handle:
    yaml.safe_dump(
      payload,
      handle,
      sort_keys=False,
      allow_unicode=False,
      indent=2,
      default_flow_style=False,
    )


def get_codepath_name(row: dict[str, str]) -> str:
  full_name = row.get("Full Name", "").strip()
  if full_name:
    return full_name
  return f"{row['First Name'].strip()} {row['Last Name'].strip()}".strip()


def build_exact_name_indexes(
  canvas_names: list[str],
) -> tuple[dict[tuple[str, ...], list[str]], dict[tuple[str, ...], list[str]]]:
  exact: dict[tuple[str, ...], list[str]] = {}
  without_initials: dict[tuple[str, ...], list[str]] = {}

  for canvas_name in canvas_names:
    exact.setdefault(token_key(canvas_name), []).append(canvas_name)
    without_initials.setdefault(token_key_without_initials(canvas_name), []).append(canvas_name)

  return exact, without_initials


def get_match_suggestions(
  codepath_name: str,
  available_canvas_names: list[str],
  suggestion_count: int,
) -> list[MatchSuggestion]:
  if not available_canvas_names:
    return []

  matches = process.extract(
    codepath_name,
    available_canvas_names,
    scorer=fuzz.token_set_ratio,
    processor=normalize_name,
    limit=suggestion_count,
  )
  return [MatchSuggestion(canvas_name=name, score=score) for name, score in matches]


def resolve_name_matches(
  codepath_names: list[str],
  canvas_names: list[str],
  existing_map: dict[str, str],
  auto_match_threshold: int,
  auto_match_gap: int,
  suggestion_count: int,
) -> tuple[dict[str, str], dict[str, str], dict[str, list[MatchSuggestion]], list[str]]:
  exact_index, no_initials_index = build_exact_name_indexes(canvas_names)
  confirmed_matches: dict[str, str] = {}
  suggested_matches: dict[str, str] = {}
  unresolved_suggestions: dict[str, list[MatchSuggestion]] = {}
  warnings: list[str] = []

  for codepath_name in codepath_names:
    mapped_canvas_name = existing_map.get(codepath_name)
    if mapped_canvas_name:
      if mapped_canvas_name not in canvas_names:
        warnings.append(
          f"Name map entry for {codepath_name!r} points to missing Canvas name {mapped_canvas_name!r}."
        )
      elif (
        mapped_canvas_name in confirmed_matches.values()
        or mapped_canvas_name in suggested_matches.values()
      ):
        warnings.append(
          f"Name map entry for {codepath_name!r} duplicates Canvas name {mapped_canvas_name!r}."
        )
      else:
        confirmed_matches[codepath_name] = mapped_canvas_name
        continue

    exact_matches = [
      candidate
      for candidate in exact_index.get(token_key(codepath_name), [])
      if candidate not in confirmed_matches.values() and candidate not in suggested_matches.values()
    ]
    if len(exact_matches) == 1:
      confirmed_matches[codepath_name] = exact_matches[0]
      continue

    no_initials_matches = [
      candidate
      for candidate in no_initials_index.get(token_key_without_initials(codepath_name), [])
      if candidate not in confirmed_matches.values() and candidate not in suggested_matches.values()
    ]
    if len(no_initials_matches) == 1:
      confirmed_matches[codepath_name] = no_initials_matches[0]
      continue

    suggestions = get_match_suggestions(
      codepath_name,
      [
        name
        for name in canvas_names
        if name not in confirmed_matches.values() and name not in suggested_matches.values()
      ],
      suggestion_count=suggestion_count,
    )
    unresolved_suggestions[codepath_name] = suggestions

    if suggestions:
      top_score = suggestions[0].score
      next_score = suggestions[1].score if len(suggestions) > 1 else -1
      if top_score >= auto_match_threshold and (top_score - next_score) >= auto_match_gap:
        suggested_matches[codepath_name] = suggestions[0].canvas_name

  return confirmed_matches, suggested_matches, unresolved_suggestions, warnings


def prompt_for_matches(
  unresolved_suggestions: dict[str, list[MatchSuggestion]],
  resolved: dict[str, str],
  canvas_names: list[str],
) -> dict[str, str]:
  updated = dict(resolved)
  used_canvas_names = set(updated.values())

  for codepath_name in sorted(unresolved_suggestions):
    candidates = [
      candidate
      for candidate in unresolved_suggestions[codepath_name]
      if candidate.canvas_name not in used_canvas_names
    ]

    print(f"\nMatch for CodePath student: {codepath_name}")
    if candidates:
      for index, candidate in enumerate(candidates, start=1):
        print(f"  {index}. {candidate.canvas_name} (score {candidate.score})")
    else:
      print("  No unused fuzzy suggestions available.")

    print("  Type a number, an exact Canvas name, 's' to skip, or 'q' to quit.")
    response = input("> ").strip()

    if not response or response.lower() == "s":
      continue
    if response.lower() == "q":
      break
    if response.isdigit():
      choice = int(response)
      if 1 <= choice <= len(candidates):
        selected = candidates[choice - 1].canvas_name
        updated[codepath_name] = selected
        used_canvas_names.add(selected)
      continue
    if response in canvas_names and response not in used_canvas_names:
      updated[codepath_name] = response
      used_canvas_names.add(response)
      continue

    print("  Ignoring invalid response.")

  return updated


def format_score(score: float | None) -> str:
  if score is None:
    return ""

  rounded = round(score, 2)
  if rounded.is_integer():
    return str(int(rounded))
  return f"{rounded:.2f}".rstrip("0").rstrip(".")


def build_score_breakdown(raw_score: float, config: ScoreConfig) -> ScoreBreakdown:
  adjusted_score = min(max(raw_score, 0.0), config.effective_total_points)
  base_earned = min(adjusted_score, config.effective_base_points)
  stretch_raw = max(adjusted_score - config.effective_base_points, 0.0)
  weighted_score = base_earned + stretch_raw * config.stretch_weight
  return ScoreBreakdown(
    raw_score=raw_score,
    adjusted_raw_score=adjusted_score,
    base_earned=base_earned,
    stretch_earned=stretch_raw,
    weighted_total=weighted_score,
    canvas_score=weighted_score * config.canvas_scale,
  )


def convert_feature_score(raw_score: float, config: ScoreConfig) -> float:
  return build_score_breakdown(raw_score, config).canvas_score


def validate_score_config(config: ScoreConfig) -> list[str]:
  warnings: list[str] = []

  if config.ignore_points > config.stretch_points:
    warnings.append(
      "--ignore-points exceeds the stretch bucket, so some base points are effectively being removed."
    )
  if config.raw_output_points == 0 and config.canvas_value > 0:
    warnings.append(
      "The converted assignment has 0 raw available points, so every output score will be 0."
    )
  if config.stretch_weight > 1:
    warnings.append("Stretch points are worth more than base points.")
  if config.stretch_weight == 0 and config.effective_stretch_points > 0:
    warnings.append("Stretch points contribute nothing with --stretch-weight 0.")
  if config.canvas_scale > 0 and config.canvas_scale != 1:
    warnings.append(
      f"Converted raw scores are being scaled by {config.canvas_scale:.4f} to fit Canvas value {config.canvas_value:g}."
    )

  return warnings


def print_score_summary(config: ScoreConfig, assignment_column: str) -> None:
  print("Score configuration:")
  print(f"  Canvas column: {assignment_column}")
  print(f"  Base points: {config.base_points:g}")
  print(f"  Stretch points: {config.stretch_points:g}")
  print(f"  Ignored points: {config.ignore_points:g}")
  print(f"  Effective base bucket: {config.effective_base_points:g}")
  print(f"  Effective stretch bucket: {config.effective_stretch_points:g}")
  print(f"  Stretch weight: {config.stretch_weight:g}")
  print(f"  Raw converted total: {config.raw_output_points:g}")
  print(f"  Canvas value: {config.canvas_value:g}")
  print(f"  Canvas scale factor: {config.canvas_scale:g}")


def compute_canvas_score(
  row: dict[str, str],
  config: ScoreConfig,
  missing_as_zero: bool,
  leave_not_graded_blank: bool,
) -> float | None:
  status = row["Status"].strip().lower()
  feature_score_text = row["Feature Score"].strip()

  if status == "not graded":
    return None if leave_not_graded_blank else 0.0

  if not feature_score_text:
    return 0.0 if missing_as_zero else None

  return convert_feature_score(float(feature_score_text), config)


def build_feedback_text(
  row: dict[str, str],
  config: ScoreConfig,
  breakdown: ScoreBreakdown | None,
  *,
  submitted_at: datetime | None = None,
  seconds_late: int | None = None,
) -> str:
  status = row["Status"].strip()
  status_lower = status.lower()
  feature_score_text = row["Feature Score"].strip() or "blank"

  lines = [
    f"CodePath status: {status}",
    f"Reported feature score: {feature_score_text}",
  ]
  if status_lower == "not graded":
    lines.append("Canvas score was forced to 0 because CodePath marked this submission as Not Graded.")
    return "\n".join(lines)
  if breakdown is None:
    lines.append("No Canvas score was pushed because the grade is blank under the current flags.")
    return "\n".join(lines)

  lines.extend([
    f"Base points: {format_score(breakdown.base_earned)}/{format_score(config.effective_base_points)}",
    f"Stretch points: {format_score(breakdown.stretch_earned)}/{format_score(config.effective_stretch_points)}",
    f"Stretch weight: {config.stretch_weight:g}",
    f"Converted raw total: {format_score(breakdown.weighted_total)}/{format_score(config.raw_output_points)}",
    f"Canvas score: {format_score(breakdown.canvas_score)}/{format_score(config.canvas_value)}",
  ])
  if config.ignore_points:
    lines.append(f"Ignored top-end CodePath points: {format_score(config.ignore_points)}")
  if submitted_at is not None:
    lines.append(f"CodePath deadline timestamp used: {submitted_at.isoformat()}")
  if seconds_late is not None:
    lines.append(f"Canvas late penalty seconds: {seconds_late}")
  return "\n".join(lines)


def normalize_canvas_datetime(value) -> datetime | None:
  if value is None:
    return None
  if isinstance(value, datetime):
    if value.tzinfo is None:
      return value.replace(tzinfo=LOS_ANGELES)
    return value
  text = str(value).strip()
  if not text:
    return None
  if text.endswith("Z"):
    text = text[:-1] + "+00:00"
  parsed = datetime.fromisoformat(text)
  if parsed.tzinfo is None:
    return parsed.replace(tzinfo=LOS_ANGELES)
  return parsed


def parse_codepath_timestamp(
  value: str,
  *,
  reference_due_at: datetime | None = None,
) -> datetime | None:
  cleaned = (value or "").strip()
  if not cleaned or cleaned == "---":
    return None

  date_part, time_part = cleaned.split(" at ", 1)
  time_part = time_part.rsplit(" ", 1)[0]
  year = reference_due_at.astimezone(LOS_ANGELES).year if reference_due_at else datetime.now(LOS_ANGELES).year
  parsed = datetime.strptime(f"{year}/{date_part} {time_part}", "%Y/%m/%d %I:%M%p")
  localized = parsed.replace(tzinfo=LOS_ANGELES)

  if reference_due_at is not None:
    local_due = reference_due_at.astimezone(LOS_ANGELES)
    if (localized - local_due).days > 180:
      localized = localized.replace(year=localized.year - 1)
    elif (local_due - localized).days > 180:
      localized = localized.replace(year=localized.year + 1)

  return localized


def get_effective_submission_time(
  row: dict[str, str],
  *,
  reference_due_at: datetime | None = None,
  strict_deadlines: bool = False,
) -> datetime | None:
  submitted_at = parse_codepath_timestamp(row.get("Submitted", ""), reference_due_at=reference_due_at)
  if submitted_at is None:
    return None
  if strict_deadlines:
    updated_at = parse_codepath_timestamp(row.get("Updated", ""), reference_due_at=reference_due_at)
    if updated_at is not None:
      return updated_at
  return submitted_at


def compute_seconds_late(
  row: dict[str, str],
  *,
  due_at,
  strict_deadlines: bool = False,
) -> tuple[datetime | None, int | None]:
  normalized_due_at = normalize_canvas_datetime(due_at)
  if normalized_due_at is None:
    return None, None

  submitted_at = get_effective_submission_time(
    row,
    reference_due_at=normalized_due_at,
    strict_deadlines=strict_deadlines,
  )
  if submitted_at is None:
    return None, None

  seconds_late = max(0, int((submitted_at - normalized_due_at).total_seconds()))
  return submitted_at, seconds_late


def resolve_canvas_submission_due_at(canvas_assignment, submission) -> datetime | None:
  if submission is not None:
    for value in (
      getattr(submission, "cached_due_date", None),
      getattr(submission, "due_at", None),
    ):
      normalized = normalize_canvas_datetime(value)
      if normalized is not None:
        return normalized

    submission_assignment = getattr(submission, "assignment", None)
    if isinstance(submission_assignment, dict):
      normalized = normalize_canvas_datetime(submission_assignment.get("due_at"))
      if normalized is not None:
        return normalized
    elif submission_assignment is not None:
      normalized = normalize_canvas_datetime(getattr(submission_assignment, "due_at", None))
      if normalized is not None:
        return normalized

  return normalize_canvas_datetime(getattr(canvas_assignment, "due_at", None))


def canvas_submission_has_content(submission) -> bool:
  if submission is None:
    return False

  submitted_at = getattr(submission, "submitted_at", None)
  if isinstance(submitted_at, str):
    if submitted_at.strip():
      return True
  elif submitted_at is not None:
    return True

  submission_type = str(getattr(submission, "submission_type", "") or "").strip().lower()
  if submission_type and submission_type not in {"none", "on_paper", "not_graded"}:
    return True

  attachments = getattr(submission, "attachments", None)
  if isinstance(attachments, list) and attachments:
    return True

  for attr_name in ("body", "url"):
    value = getattr(submission, attr_name, None)
    if isinstance(value, str) and value.strip():
      return True

  media_comment_id = getattr(submission, "media_comment_id", None)
  if media_comment_id not in (None, "", 0):
    return True

  return False


def canvas_submission_is_excused(submission) -> bool:
  return bool(getattr(submission, "excused", False)) if submission is not None else False


def canvas_submission_is_missing_candidate(
  submission,
  *,
  due_at: datetime | None,
  now: datetime,
) -> bool:
  if submission is None:
    return False
  if due_at is None or due_at > now:
    return False
  if canvas_submission_is_excused(submission):
    return False
  return not canvas_submission_has_content(submission)


def decide_submission_action(
  *,
  codepath_row: dict[str, str] | None,
  canvas_assignment,
  submission,
  strict_deadlines: bool,
  now: datetime,
) -> SubmissionDecision:
  due_at = resolve_canvas_submission_due_at(canvas_assignment, submission)
  if codepath_row is not None:
    submitted_at, seconds_late = compute_seconds_late(
      codepath_row,
      due_at=due_at,
      strict_deadlines=strict_deadlines,
    )
    if submitted_at is not None:
      return SubmissionDecision(
        action="push_grade",
        due_at=due_at,
        submitted_at=submitted_at,
        seconds_late=seconds_late,
      )
    if canvas_submission_is_missing_candidate(submission, due_at=due_at, now=now):
      return SubmissionDecision(action="mark_missing", due_at=due_at)
    return SubmissionDecision(action="skip", due_at=due_at)

  if canvas_submission_is_missing_candidate(submission, due_at=due_at, now=now):
    return SubmissionDecision(action="mark_missing", due_at=due_at)
  return SubmissionDecision(action="skip", due_at=due_at)


def write_suggestions(path: Path, suggestions: dict[str, list[MatchSuggestion]]) -> None:
  payload = {
    "suggestions": {
      codepath_name: [
        {"canvas_name": candidate.canvas_name, "score": candidate.score}
        for candidate in candidates
      ]
      for codepath_name, candidates in sorted(suggestions.items())
    }
  }
  with path.open("w", encoding="utf-8") as handle:
    yaml.safe_dump(payload, handle, sort_keys=False, allow_unicode=False)


def write_canvas_output(
  path: Path,
  fieldnames: list[str],
  rows: list[dict[str, str]],
) -> None:
  with path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)


def get_name_map_cache(args: argparse.Namespace) -> dict[str, str]:
  cache = getattr(args, "_name_map_cache", None)
  if cache is None:
    name_map_path = Path(args.name_map) if getattr(args, "name_map", None) else None
    cache = load_name_map(name_map_path)
    setattr(args, "_name_map_cache", cache)
  return cache


def get_codepath_fieldnames(rows: list[dict[str, str]]) -> list[str]:
  if not rows:
    return []
  return list(rows[0].keys())


def classify_codepath_export(rows: list[dict[str, str]]) -> tuple[bool, str | None]:
  fieldnames = set(get_codepath_fieldnames(rows))
  required_fields = {"First Name", "Last Name", "Feature Score", "Status"}
  gradebook_fields = {"Full Name", "Feature Score", "Status"}
  if required_fields.issubset(fieldnames):
    return True, None
  if gradebook_fields.issubset(fieldnames):
    return True, None

  if fieldnames == {"Full Name"} or fieldnames == {"Full Name", None}:
    return False, "CodePath export only contains a roster and no grading columns yet."

  if "Full Name" in fieldnames:
    missing = sorted(gradebook_fields - fieldnames)
  else:
    missing = sorted(required_fields - fieldnames)
  return False, f"CodePath export is missing required columns: {', '.join(missing)}"


def run_single_conversion(
  args: argparse.Namespace,
  codepath_path: Path,
  canvas_path: Path,
  output_path: Path,
  assignment_name: str | None = None,
) -> int:
  name_map_path = Path(args.name_map) if args.name_map else None
  suggestions_path = Path(args.write_suggestions) if args.write_suggestions else None

  if assignment_name:
    print(f"\n=== {assignment_name} ===")
  codepath_rows = read_codepath_rows(codepath_path)
  codepath_valid, codepath_message = classify_codepath_export(codepath_rows)
  if not codepath_valid:
    if assignment_name and codepath_message == "CodePath export only contains a roster and no grading columns yet.":
      print(f"Skipping {assignment_name}: {codepath_message}")
      return 0
    print(f"Cannot process {codepath_path}: {codepath_message}", file=sys.stderr)
    return 1

  fieldnames, canvas_rows = read_canvas_rows(canvas_path)

  if not fieldnames:
    raise ValueError(f"No headers found in {canvas_path}.")

  assignment_column = args.assignment_column or fieldnames[-1]
  if assignment_column not in fieldnames:
    raise ValueError(f"Canvas column {assignment_column!r} was not found.")
  config = build_score_config(args, canvas_rows, assignment_column)
  print_score_summary(config, assignment_column)
  for warning in validate_score_config(config):
    print(f"Warning: {warning}", file=sys.stderr)

  canvas_students = [
    row["Student"]
    for row in canvas_rows
    if row.get("Student", "").strip() and row["Student"] != POINTS_POSSIBLE_LABEL
  ]
  codepath_names = [get_codepath_name(row) for row in codepath_rows]

  existing_map = dict(get_name_map_cache(args))
  confirmed_matches, suggested_matches, unresolved_suggestions, warnings = resolve_name_matches(
    codepath_names=codepath_names,
    canvas_names=canvas_students,
    existing_map=existing_map,
    auto_match_threshold=args.auto_match_threshold,
    auto_match_gap=args.auto_match_gap,
    suggestion_count=args.suggestion_count,
  )
  resolved_matches = dict(confirmed_matches)

  if args.prompt_for_matches:
    if not sys.stdin.isatty():
      raise ValueError("--prompt-for-matches requires an interactive terminal.")
    resolved_matches = prompt_for_matches(unresolved_suggestions, resolved_matches, canvas_students)
    confirmed_matches = dict(resolved_matches)
    unresolved_suggestions = {
      codepath_name: candidates
      for codepath_name, candidates in unresolved_suggestions.items()
      if codepath_name not in resolved_matches
    }
    suggested_matches = {
      codepath_name: canvas_name
      for codepath_name, canvas_name in suggested_matches.items()
      if codepath_name not in resolved_matches
    }

  if name_map_path is not None:
    merged_confirmed_map = {
      codepath_name: canvas_name
      for codepath_name, canvas_name in existing_map.items()
      if codepath_name not in codepath_names
    }
    merged_confirmed_map.update(confirmed_matches)
    suggested_name_map = dict(suggested_matches)
    for codepath_name, suggestions in unresolved_suggestions.items():
      if suggestions:
        suggested_name_map.setdefault(codepath_name, suggestions[0].canvas_name)
    save_name_map(
      name_map_path,
      merged_confirmed_map,
      suggested_mapping=suggested_name_map,
      unmatched=sorted(name for name, suggestions in unresolved_suggestions.items() if not suggestions),
    )
  shared_name_map = get_name_map_cache(args)
  shared_name_map.clear()
  shared_name_map.update(merged_confirmed_map)

  if suggestions_path is not None:
    write_suggestions(suggestions_path, unresolved_suggestions)

  if suggested_matches or unresolved_suggestions:
    unresolved_without_suggestions = [
      codepath_name
      for codepath_name, suggestions in unresolved_suggestions.items()
      if not suggestions
    ]
    print(
      "Refusing to write Canvas grades until all fuzzy suggestions are manually confirmed.",
      file=sys.stderr,
    )
    if suggested_matches:
      print(
        f"Review required for {len(suggested_matches)} suggested match(es) in the name map.",
        file=sys.stderr,
      )
    if unresolved_without_suggestions:
      print(
        f"No suggestions available for {len(unresolved_without_suggestions)} unmatched student(s).",
        file=sys.stderr,
      )
    print(
      "Re-run with --prompt-for-matches or move reviewed entries from SUGGESTED to CONFIRMED.",
      file=sys.stderr,
    )
    return 1

  codepath_by_name = {get_codepath_name(row): row for row in codepath_rows}
  codepath_by_canvas_name = {
    canvas_name: codepath_name for codepath_name, canvas_name in resolved_matches.items()
  }

  for row in canvas_rows:
    student = row.get("Student", "")
    if student == POINTS_POSSIBLE_LABEL:
      row[assignment_column] = format_score(config.canvas_value)
      continue
    if student not in codepath_by_canvas_name:
      continue

    codepath_name = codepath_by_canvas_name[student]
    codepath_row = codepath_by_name[codepath_name]
    row[assignment_column] = format_score(
      compute_canvas_score(
        codepath_row,
        config=config,
        missing_as_zero=args.missing_as_zero,
        leave_not_graded_blank=args.leave_not_graded_blank,
      )
    )

  write_canvas_output(output_path, fieldnames, canvas_rows)

  matched_canvas_names = set(resolved_matches.values())
  unmatched_canvas = sorted(set(canvas_students) - matched_canvas_names)
  unmatched_codepath = sorted(set(codepath_names) - set(resolved_matches))

  for warning in warnings:
    print(f"Warning: {warning}", file=sys.stderr)

  print(f"Wrote {output_path}")
  print(f"Matched {len(resolved_matches)} of {len(codepath_rows)} CodePath students")
  print(f"Unmatched CodePath students: {len(unmatched_codepath)}")
  print(f"Canvas students left blank: {len(unmatched_canvas)}")
  if unresolved_suggestions:
    print("Unresolved suggestions:")
    for codepath_name, suggestions in sorted(unresolved_suggestions.items()):
      summary = ", ".join(f"{candidate.canvas_name} ({candidate.score})" for candidate in suggestions[:3])
      print(f"  {codepath_name}: {summary}")

  return 0


def load_assignments_config(path: Path) -> tuple[int | None, dict[str, dict[str, object]]]:
  loaded = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
  if not isinstance(loaded, dict):
    raise ValueError(f"Assignments config at {path} must be a mapping.")

  course_id = loaded.get("course-id", loaded.get("course_id"))
  assignments = loaded.get("assignments")
  if assignments is None:
    assignments = {
      name: settings
      for name, settings in loaded.items()
      if isinstance(settings, dict)
    }

  if not isinstance(assignments, dict):
    raise ValueError(f"Assignments config at {path} must contain an assignments mapping.")

  for assignment_name, settings in assignments.items():
    if not isinstance(settings, dict):
      raise ValueError(f"Assignment {assignment_name!r} must map to a settings dictionary.")
  return course_id, assignments


def build_assignment_args(
  base_args: argparse.Namespace,
  assignment_name: str,
  settings: dict[str, object],
) -> argparse.Namespace:
  assignment_args = argparse.Namespace(**vars(base_args))
  assignment_args.base_points = settings.get("base", settings.get("base_points", base_args.base_points))
  assignment_args.stretch_points = settings.get(
    "stretch",
    settings.get("stretch_points", base_args.stretch_points),
  )
  assignment_args.ignore_points = settings.get(
    "ignore",
    settings.get("ignore_points", base_args.ignore_points),
  )
  assignment_args.stretch_weight = settings.get(
    "stretch_weight",
    settings.get("stretch-weight", base_args.stretch_weight),
  )
  assignment_args.canvas_value = settings.get("canvas_value", base_args.canvas_value)

  if assignment_args.base_points is None or assignment_args.stretch_points is None:
    raise ValueError(
      f"Assignment {assignment_name!r} is missing base/stretch settings in the YAML."
    )

  return assignment_args


def get_canvas_roster_rows_from_course(course) -> list[dict[str, str]]:
  return [
    {
      "Student": student.name,
      "ID": str(student.user_id),
    }
    for student in course.get_students(include_names=True)
  ]


def push_feedback_accepts_seconds_late(canvas_assignment) -> bool:
  try:
    signature = inspect.signature(canvas_assignment.push_feedback)
  except (TypeError, ValueError):
    return False
  if "seconds_late" in signature.parameters:
    return True
  return any(
    parameter.kind is inspect.Parameter.VAR_KEYWORD
    for parameter in signature.parameters.values()
  )


def mark_canvas_submission_missing(canvas_assignment, user_id: int) -> bool:
  try:
    submission = canvas_assignment.get_submission(user_id)
    submission.edit(submission={"late_policy_status": "missing"})
  except Exception:
    return False
  return True


def run_single_push_conversion(
  args: argparse.Namespace,
  codepath_path: Path | None,
  roster_rows: list[dict[str, str]],
  canvas_assignment,
  assignment_name: str,
  *,
  codepath_rows: list[dict[str, str]] | None = None,
  push_enabled: bool = True,
) -> int:
  print(f"\n=== {assignment_name} ===")
  if codepath_rows is None:
    if codepath_path is None:
      raise ValueError("run_single_push_conversion requires either codepath_path or codepath_rows.")
    codepath_rows = read_codepath_rows(codepath_path)
  codepath_valid, codepath_message = classify_codepath_export(codepath_rows)
  if not codepath_valid:
    if codepath_message == "CodePath export only contains a roster and no grading columns yet.":
      print(f"Skipping {assignment_name}: {codepath_message}")
      return 0
    source_label = str(codepath_path) if codepath_path is not None else assignment_name
    print(f"Cannot process {source_label}: {codepath_message}", file=sys.stderr)
    return 1

  args = argparse.Namespace(**vars(args))
  args.canvas_value = getattr(canvas_assignment, "points_possible", None)
  config = build_score_config(args, [], f"Canvas assignment {canvas_assignment.id}")
  print_score_summary(config, f"{canvas_assignment.name} ({canvas_assignment.id})")
  for warning in validate_score_config(config):
    print(f"Warning: {warning}", file=sys.stderr)

  canvas_students = [row["Student"] for row in roster_rows]
  codepath_names = [get_codepath_name(row) for row in codepath_rows]
  name_map_path = Path(args.name_map) if args.name_map else None
  suggestions_path = Path(args.write_suggestions) if args.write_suggestions else None

  existing_map = dict(get_name_map_cache(args))
  confirmed_matches, suggested_matches, unresolved_suggestions, warnings = resolve_name_matches(
    codepath_names=codepath_names,
    canvas_names=canvas_students,
    existing_map=existing_map,
    auto_match_threshold=args.auto_match_threshold,
    auto_match_gap=args.auto_match_gap,
    suggestion_count=args.suggestion_count,
  )
  resolved_matches = dict(confirmed_matches)

  if args.prompt_for_matches:
    if not sys.stdin.isatty():
      raise ValueError("--prompt-for-matches requires an interactive terminal.")
    resolved_matches = prompt_for_matches(unresolved_suggestions, resolved_matches, canvas_students)
    confirmed_matches = dict(resolved_matches)
    unresolved_suggestions = {
      codepath_name: candidates
      for codepath_name, candidates in unresolved_suggestions.items()
      if codepath_name not in resolved_matches
    }
    suggested_matches = {
      codepath_name: canvas_name
      for codepath_name, canvas_name in suggested_matches.items()
      if codepath_name not in resolved_matches
    }

  if name_map_path is not None:
    merged_confirmed_map = {
      codepath_name: canvas_name
      for codepath_name, canvas_name in existing_map.items()
      if codepath_name not in codepath_names
    }
    merged_confirmed_map.update(confirmed_matches)
    suggested_name_map = dict(suggested_matches)
    for codepath_name, suggestions in unresolved_suggestions.items():
      if suggestions:
        suggested_name_map.setdefault(codepath_name, suggestions[0].canvas_name)
    save_name_map(
      name_map_path,
      merged_confirmed_map,
      suggested_mapping=suggested_name_map,
      unmatched=sorted(name for name, suggestions in unresolved_suggestions.items() if not suggestions),
    )
  shared_name_map = get_name_map_cache(args)
  shared_name_map.clear()
  shared_name_map.update(merged_confirmed_map)

  if suggestions_path is not None:
    write_suggestions(suggestions_path, unresolved_suggestions)

  if suggested_matches or unresolved_suggestions:
    unresolved_without_suggestions = [
      codepath_name
      for codepath_name, suggestions in unresolved_suggestions.items()
      if not suggestions
    ]
    print(
      "Refusing to push Canvas grades until all fuzzy suggestions are manually confirmed.",
      file=sys.stderr,
    )
    if suggested_matches:
      print(
        f"Review required for {len(suggested_matches)} suggested match(es) in the name map.",
        file=sys.stderr,
      )
      for codepath_name in sorted(suggested_matches):
        suggestions = unresolved_suggestions.get(codepath_name, [])
        top_suggestion = suggestions[0] if suggestions else None
        if top_suggestion is None:
          print(f"  {codepath_name} -> {suggested_matches[codepath_name]}", file=sys.stderr)
        else:
          print(
            f"  {codepath_name} -> {top_suggestion.canvas_name} ({top_suggestion.score})",
            file=sys.stderr,
          )
    if unresolved_without_suggestions:
      print(
        f"No suggestions available for {len(unresolved_without_suggestions)} unmatched student(s).",
        file=sys.stderr,
      )
      for codepath_name in unresolved_without_suggestions:
        print(f"  {codepath_name}", file=sys.stderr)
    print(
      "Re-run with --prompt-for-matches or move reviewed entries from SUGGESTED to CONFIRMED.",
      file=sys.stderr,
    )
    return 1

  matched_canvas_names = set(resolved_matches.values())
  unmatched_canvas_names = sorted(set(canvas_students) - matched_canvas_names)
  if unmatched_canvas_names:
    print(
      f"Warning: {len(unmatched_canvas_names)} Canvas roster student(s) have no CodePath match; skipping them.",
      file=sys.stderr,
    )
    if args.verbose:
      for canvas_name in unmatched_canvas_names:
        print(f"  {canvas_name}", file=sys.stderr)
    elif len(unmatched_canvas_names) <= 10:
      for canvas_name in unmatched_canvas_names:
        print(f"  {canvas_name}", file=sys.stderr)
    else:
      for canvas_name in unmatched_canvas_names[:10]:
        print(f"  {canvas_name}", file=sys.stderr)
      print("  (use --verbose to list all unmatched Canvas students)", file=sys.stderr)

  if not push_enabled:
    print(f"Preflight OK for {assignment_name}: {len(resolved_matches)} matched student(s)")
    return 0

  codepath_by_name = {get_codepath_name(row): row for row in codepath_rows}
  codepath_by_canvas_name = {
    canvas_name: codepath_by_name[codepath_name]
    for codepath_name, canvas_name in resolved_matches.items()
  }
  pushed_count = 0
  marked_missing_count = 0
  skipped_blank_count = 0
  skipped_no_action_count = 0
  skipped_unmatched_canvas_count = 0
  failed_push_count = 0
  now = datetime.now(LOS_ANGELES)

  for roster_row in roster_rows:
    canvas_name = roster_row["Student"]
    user_id_text = str(roster_row.get("ID", "")).strip()
    if not user_id_text:
      print(f"Missing Canvas user ID for {canvas_name}.", file=sys.stderr)
      failed_push_count += 1
      continue

    user_id = int(user_id_text)
    try:
      submission = canvas_assignment.get_submission(user_id)
    except Exception:
      print(f"Could not fetch Canvas submission for {canvas_name}.", file=sys.stderr)
      failed_push_count += 1
      continue

    codepath_row = codepath_by_canvas_name.get(canvas_name)
    decision = decide_submission_action(
      codepath_row=codepath_row,
      canvas_assignment=canvas_assignment,
      submission=submission,
      strict_deadlines=args.strict_deadlines,
      now=now,
    )

    if decision.action == "mark_missing":
      if mark_canvas_submission_missing(canvas_assignment, user_id):
        marked_missing_count += 1
      else:
        failed_push_count += 1
      continue
    if decision.action == "skip":
      if codepath_row is None:
        skipped_unmatched_canvas_count += 1
        continue
      skipped_no_action_count += 1
      continue
    if codepath_row is None:
      skipped_no_action_count += 1
      continue

    score = compute_canvas_score(
      codepath_row,
      config=config,
      missing_as_zero=args.missing_as_zero,
      leave_not_graded_blank=args.leave_not_graded_blank,
    )
    if score is None:
      skipped_blank_count += 1
      continue

    breakdown = build_score_breakdown(float(codepath_row["Feature Score"]), config)
    feedback_text = build_feedback_text(
      codepath_row,
      config,
      breakdown,
      submitted_at=decision.submitted_at,
      seconds_late=decision.seconds_late,
    )
    push_kwargs = {
      "user_id": user_id,
      "score": score,
      "comments": feedback_text,
      "keep_previous_best": True,
      "clobber_feedback": False,
    }
    if decision.seconds_late is not None and push_feedback_accepts_seconds_late(canvas_assignment):
      push_kwargs["seconds_late"] = decision.seconds_late
    pushed = canvas_assignment.push_feedback(
      **push_kwargs,
    )
    if pushed:
      pushed_count += 1
    else:
      failed_push_count += 1

  for warning in warnings:
    print(f"Warning: {warning}", file=sys.stderr)

  print(f"Pushed {pushed_count} grade(s) to Canvas for {assignment_name}")
  print(f"Marked missing submissions: {marked_missing_count}")
  print(f"Skipped blank scores: {skipped_blank_count}")
  print(f"Skipped with no action: {skipped_no_action_count}")
  print(f"Skipped unmatched Canvas roster students: {skipped_unmatched_canvas_count}")
  print(f"Push failures: {failed_push_count}")
  return 1 if failed_push_count else 0


def run_batch_conversion(args: argparse.Namespace) -> int:
  course_id, assignments = load_assignments_config(Path(args.assignments))
  selected_assignments = set(args.only_assignment or assignments.keys())
  data_dir = Path(args.data_dir) if args.data_dir else Path.cwd()
  explicit_xls = Path(args.xls) if args.xls else None
  gradebook_path = explicit_xls or find_gradebook_workbook(data_dir)
  exit_code = 0
  push_mode = course_id is not None
  course = None
  roster_rows: list[dict[str, str]] | None = None
  args._name_map_cache = load_name_map(Path(args.name_map)) if args.name_map else {}

  if explicit_xls is not None:
    if not explicit_xls.exists():
      raise ValueError(f"Workbook not found: {explicit_xls}")
    if explicit_xls.suffix.lower() != ".xlsx":
      raise ValueError(f"Workbook must be an .xlsx file: {explicit_xls}")
  if explicit_xls is not None and not push_mode:
    raise ValueError("--xls is only supported for Canvas push mode right now.")

  if push_mode:
    canvas_interface = CanvasInterface(prod=args.prod, privacy_mode="none")
    course = canvas_interface.get_course(int(course_id))
    roster_rows = get_canvas_roster_rows_from_course(course)
    print(f"Canvas target: {'PROD' if args.prod else 'DEV'}")
    if explicit_xls is not None:
      print(f"Workbook source: {explicit_xls}")

    preflight_failed = False
    for assignment_name, settings in assignments.items():
      if assignment_name not in selected_assignments:
        continue
      print(f"Preflighting {assignment_name}...", flush=True)

      codepath_path, codepath_rows, skip_message, missing_error = resolve_batch_assignment_input(
        assignment_name=assignment_name,
        data_dir=data_dir,
        gradebook_path=gradebook_path,
        prefer_gradebook=explicit_xls is not None,
      )
      if skip_message is not None:
        print(f"Skipping {assignment_name}: {skip_message}")
        continue
      if missing_error is not None:
        print(missing_error, file=sys.stderr)
        exit_code = 1
        preflight_failed = True
        continue

      assignment_args = build_assignment_args(args, assignment_name, settings)
      assignment_id = settings.get("assignment-id", settings.get("assignment_id"))
      if assignment_id is None:
        print(f"Missing assignment-id for {assignment_name} in assignments config.", file=sys.stderr)
        exit_code = 1
        preflight_failed = True
        continue
      canvas_assignment = course.get_assignment(int(assignment_id))
      if canvas_assignment is None:
        print(f"Could not find Canvas assignment {assignment_id} for {assignment_name}.", file=sys.stderr)
        exit_code = 1
        preflight_failed = True
        continue
      result = run_single_push_conversion(
        assignment_args,
        codepath_path=codepath_path,
        roster_rows=roster_rows or [],
        canvas_assignment=canvas_assignment,
        assignment_name=assignment_name,
        codepath_rows=codepath_rows,
        push_enabled=False,
      )
      if result != 0:
        exit_code = 1
        preflight_failed = True

    if preflight_failed:
      print("No Canvas grades were pushed because preflight found unresolved matches.", file=sys.stderr)
      return exit_code

  for assignment_name, settings in assignments.items():
    if assignment_name not in selected_assignments:
      continue
    if push_mode:
      print(f"Pushing {assignment_name}...", flush=True)

    codepath_path, codepath_rows, skip_message, missing_error = resolve_batch_assignment_input(
      assignment_name=assignment_name,
      data_dir=data_dir,
      gradebook_path=gradebook_path if push_mode else None,
      prefer_gradebook=explicit_xls is not None,
    )
    if skip_message is not None:
      print(f"Skipping {assignment_name}: {skip_message}")
      continue
    if missing_error is not None:
      print(missing_error, file=sys.stderr)
      exit_code = 1
      continue

    canvas_path = data_dir / f"canvas-{assignment_name}.csv"

    assignment_args = build_assignment_args(args, assignment_name, settings)
    if push_mode:
      assignment_id = settings.get("assignment-id", settings.get("assignment_id"))
      canvas_assignment = course.get_assignment(int(assignment_id))
      result = run_single_push_conversion(
        assignment_args,
        codepath_path=codepath_path,
        roster_rows=roster_rows or [],
        canvas_assignment=canvas_assignment,
        assignment_name=assignment_name,
        codepath_rows=codepath_rows,
        push_enabled=True,
      )
    else:
      if codepath_path is None:
        print(
          f"Workbook-backed batch input is only supported for Canvas push mode right now ({assignment_name}).",
          file=sys.stderr,
        )
        exit_code = 1
        continue
      if not canvas_path.exists():
        print(
          f"Missing Canvas CSV for {assignment_name}: expected {canvas_path.name}",
          file=sys.stderr,
        )
        exit_code = 1
        continue
      result = run_single_conversion(
        assignment_args,
        codepath_path=codepath_path,
        canvas_path=canvas_path,
        output_path=canvas_path,
        assignment_name=assignment_name,
      )
    if result != 0:
      exit_code = 1

  return exit_code


def main(argv: list[str] | None = None) -> int:
  args = parse_args(argv)
  args._name_map_cache = load_name_map(Path(args.name_map)) if args.name_map else {}
  if args.assignments:
    return run_batch_conversion(args)

  codepath_path = Path(args.codepath_csv)
  canvas_path = Path(args.canvas)
  output_path = Path(args.out) if args.out else canvas_path
  return run_single_conversion(
    args,
    codepath_path=codepath_path,
    canvas_path=canvas_path,
    output_path=output_path,
  )


if __name__ == "__main__":
  raise SystemExit(main())
